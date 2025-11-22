#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to HTML Converter for Prefecture Subsidy Data
都道府県別補助金データ エクセル→HTML変換ツール

使い方 (Usage):
    python excel_to_html_converter.py <excel_file> <output_file>
    
    例 (Example):
    python excel_to_html_converter.py koufu_data.xlsx koufu.html
    python excel_to_html_converter.py koufu2_data.xlsx koufu2.html
    python excel_to_html_converter.py koufu3_data.xlsx koufu3.html

エクセルファイル形式 (Excel Format):
    - 列A: 都道府県名 (Prefecture names)
    - 列B以降: 各年度のデータ (Year data: 2009, 2010, 2011...)
    - 最終行: 総計 (Total row with "総計" in column A)
    - 複数シート対応: 各タブ名がシート名 (Multiple sheets: each sheet name becomes a tab)
    
    例 (Example):
    | 都道府県 | 2009 | 2010 | 2011 | ... | 計 |
    |---------|------|------|------|-----|-----|
    | 北海道   | 596  | 654  | 560  | ... | 11559 |
    | 青森県   | 192  | 261  | 159  | ... | 3015 |
    | ...     | ...  | ...  | ...  | ... | ... |
    | 総計    | 1234 | 5678 | 9012 | ... | 99999 |
"""

import sys
import openpyxl
from datetime import datetime
from pathlib import Path
import platform


def format_japanese_date(dt=None):
    """
    日付を日本語形式でフォーマット（ゼロ埋めなし）
    Format date in Japanese style without zero-padding
    Cross-platform compatible
    
    Args:
        dt: datetime object (default: current datetime)
    
    Returns:
        str: Formatted date string (e.g., "2025年11月21日")
    """
    if dt is None:
        dt = datetime.now()
    
    # Platform-specific format for non-zero-padded dates
    if platform.system() == 'Windows':
        # Windows uses %# for non-zero-padded
        try:
            return dt.strftime('%Y年%#m月%#d日')
        except (ValueError, AttributeError):
            # Fallback to manual zero removal
            return dt.strftime('%Y年%m月%d日').replace('年0', '年').replace('月0', '月')
    else:
        # Unix-like systems use %-
        try:
            return dt.strftime('%Y年%-m月%-d日')
        except (ValueError, AttributeError):
            # Fallback to manual zero removal
            return dt.strftime('%Y年%m月%d日').replace('年0', '年').replace('月0', '月')


def format_number(value):
    """
    数値を3桁ごとのカンマ区切りでフォーマット
    Format number with thousand separators
    
    Args:
        value: The value to format (can be string or number)
    
    Returns:
        str: Formatted string with commas or original value
    """
    if value == '－' or value == '' or value is None:
        return '－'
    
    try:
        # 数値に変換してみる
        num = float(str(value).replace(',', ''))
        # 整数かチェック
        if num.is_integer():
            return f"{int(num):,}"
        else:
            return f"{num:,.1f}"
    except (ValueError, AttributeError):
        # 数値でない場合はそのまま返す
        return str(value)


def read_excel_data(excel_file):
    """
    エクセルファイルからデータを読み込む
    Read data from Excel file
    
    Returns:
        dict: {sheet_name: {'years': [...], 'prefectures': {...}}}
    """
    wb = openpyxl.load_workbook(excel_file)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # ヘッダー行から年度を取得 (Get years from header row)
        years = []
        for col in range(2, ws.max_column + 1):
            year_value = ws.cell(1, col).value
            if year_value:
                years.append(str(year_value))
        
        # 都道府県データを取得 (Get prefecture data)
        prefectures = {}
        total_row = None
        
        for row in range(2, ws.max_row + 1):
            pref_name = ws.cell(row, 1).value
            if not pref_name:
                continue
                
            # データ行を取得 (Get data row)
            row_data = []
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                # 数値またはハイフン (Number or dash)
                if cell_value is None or cell_value == '':
                    row_data.append('－')
                else:
                    # 数値をフォーマット (Format numbers with commas)
                    row_data.append(format_number(cell_value))
            
            # 総計行かチェック (Check if total row)
            if '総計' in str(pref_name) or '合計' in str(pref_name):
                total_row = row_data
            else:
                prefectures[pref_name] = row_data
        
        data[sheet_name] = {
            'years': years,
            'prefectures': prefectures,
            'total': total_row
        }
    
    wb.close()
    return data


def generate_html(data, output_file, page_title, page_subtitle):
    """
    HTMLファイルを生成 (Generate HTML file)
    """
    
    # タブのリスト (List of tabs)
    tabs = list(data.keys())
    
    html_template = f'''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{page_title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Hiragino Sans', 'Hiragino Kaku Gothic ProN', 'Noto Sans JP', 'Yu Gothic', 'Meiryo', sans-serif;
            line-height: 1.6;
            color: #333;
            background-color: #f5f5f5;
            padding: 20px;
            scroll-behavior: smooth;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        
        .main-nav {{
            text-align: center;
            padding: 5px 0;
            margin-bottom: 15px;
            font-size: 11px;
            color: #666;
        }}
        
        .main-nav a {{
            color: #666;
            text-decoration: none;
            margin: 0 10px;
        }}
        
        .main-nav a:hover {{
            color: #0066cc;
            text-decoration: underline;
        }}
        
        h1 {{
            color: #0066cc;
            margin-bottom: 10px;
            font-size: 28px;
            text-align: left;
        }}
        
        .subtitle {{
            text-align: left;
            color: #666;
            margin-bottom: 20px;
            font-size: 16px;
        }}
        
        .cross-nav {{
            text-align: left;
            padding: 10px;
            background-color: #f0f0f0;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        
        .cross-nav a {{
            color: #0066cc;
            text-decoration: none;
            margin: 0 15px;
            font-size: 14px;
        }}
        
        .cross-nav a:hover {{
            text-decoration: underline;
        }}
        
        .info-section {{
            background-color: #e8f4f8;
            padding: 15px;
            margin-bottom: 20px;
            border-left: 4px solid #0066cc;
            font-size: 14px;
        }}
        
        .info-section p {{
            margin: 5px 0;
        }}
        
        .tabs {{
            display: flex;
            gap: 5px;
            margin-bottom: 20px;
            border-bottom: 2px solid #0066cc;
        }}
        
        .tab {{
            padding: 10px 20px;
            background-color: #f0f0f0;
            border: none;
            cursor: pointer;
            font-size: 16px;
            transition: all 0.3s;
            border-radius: 4px 4px 0 0;
        }}
        
        .tab:hover {{
            background-color: #e0e0e0;
        }}
        
        .tab.active {{
            background-color: #0066cc;
            color: white;
        }}
        
        .tab-content {{
            display: none;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        .table-wrapper {{
            overflow-x: auto;
            overflow-y: auto;
            max-height: 600px;
            margin-top: 20px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}
        
        th, td {{
            padding: 12px 8px;
            text-align: right;
            border: 1px solid #ddd;
        }}
        
        th {{
            background-color: #0066cc;
            color: white;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
            text-align: center;
        }}
        
        th:first-child, td:first-child {{
            text-align: left;
            position: sticky;
            left: 0;
            background-color: white;
            z-index: 5;
        }}
        
        th:first-child {{
            z-index: 15;
            background-color: #0066cc;
        }}
        
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        
        tr:hover {{
            background-color: #f0f8ff;
        }}
        
        /* 総計行のスタイル (Total row style) */
        tr.total-row {{
            background-color: #d4edda !important;
            font-weight: bold;
            border-top: 2px solid #28a745;
            border-bottom: 2px solid #28a745;
        }}
        
        tr.total-row td {{
            color: #000;
        }}
        
        /* 計列のスタイル (Total column style) */
        th.total-col {{
            background-color: #0066cc;
            color: white;
            border-left: 2px solid #0066cc;
            border-right: 2px solid #0066cc;
        }}
        
        td.total-col {{
            background-color: #cce5ff;
            font-weight: bold;
            border-left: 2px solid #0066cc;
            border-right: 2px solid #0066cc;
        }}
        
        /* 総計行と計列の交差セル (Intersection cell) */
        tr.total-row td.total-col {{
            background-color: #0066cc;
            color: white;
        }}
        
        @media (max-width: 768px) {{
            body {{
                padding: 10px;
            }}
            
            .container {{
                padding: 15px;
            }}
            
            h1 {{
                font-size: 22px;
            }}
            
            table {{
                font-size: 12px;
            }}
            
            th, td {{
                padding: 8px 4px;
            }}
        }}
        
        @media (max-width: 480px) {{
            table {{
                font-size: 11px;
            }}
            
            th, td {{
                padding: 6px 3px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <!-- メインナビゲーション (Main Navigation) -->
        <div class="main-nav">
            <a href="https://www.cev-pc.or.jp/">HOME</a> | 
            <a href="/tokei/hoyuudaisu.html">EV等 保有台数統計</a> | 
            <a href="/tokei/hanbaidaisu.html">EV等 販売台数統計</a>
        </div>
        
        <h1>{page_title}</h1>
        <div class="subtitle">{page_subtitle}</div>
        
        <!-- クロスページナビゲーション (Cross-page Navigation) -->
        <div class="cross-nav" id="cross-nav">
            <!-- Will be filled by JavaScript based on current page -->
        </div>
        
        <div class="info-section">
            <p>○ {format_japanese_date()} 次世代自動車振興センター</p>
            <p>○ {datetime.now().strftime('%Y年度は')} {format_japanese_date()} までの集計です</p>
            <p>※{datetime.now().strftime('%Y年度')}の補助金交付台数等については、現在審査中のものもあるため、{format_japanese_date()}現在の数値であり、第6次公募締切（予定）までの最終的な数値ではありません。</p>
            <p>※ここで使用されている数字について</p>
            <p>※※FCV（燃料電池自動車）の交付台数は2014年からの集計です</p>
            <p>※※外部給電器と原付EVの交付台数は2020年からの集計です</p>
            <p>※※V2H充放電設備の交付基数は2020年からの集計です</p>
        </div>
        
        <!-- タブ (Tabs) -->
        <div class="tabs">
'''
    
    # タブボタンを生成 (Generate tab buttons)
    for i, tab_name in enumerate(tabs):
        active_class = 'active' if i == 0 else ''
        html_template += f'            <button class="tab {active_class}" onclick="showTab(event, \'{tab_name}\')">{tab_name}</button>\n'
    
    html_template += '        </div>\n\n'
    
    # 各タブのコンテンツを生成 (Generate content for each tab)
    for i, (tab_name, tab_data) in enumerate(data.items()):
        active_class = 'active' if i == 0 else ''
        html_template += f'        <!-- {tab_name}タブ ({tab_name} Tab) -->\n'
        html_template += f'        <div id="{tab_name}" class="tab-content {active_class}">\n'
        html_template += f'            <h2>{tab_name} 都道府県別補助金交付台数一覧表（{tab_data["years"][0]}～{tab_data["years"][-1]}年度）</h2>\n'
        html_template += '            <div class="table-wrapper">\n'
        html_template += '                <table>\n'
        html_template += '                    <thead>\n'
        html_template += '                        <tr>\n'
        html_template += '                            <th>都道府県</th>\n'
        
        # 年度ヘッダー (Year headers)
        for year in tab_data['years']:
            if year == '計' or '計' in str(year):
                html_template += f'                            <th class="total-col">{year}</th>\n'
            else:
                html_template += f'                            <th>{year}</th>\n'
        
        html_template += '                        </tr>\n'
        html_template += '                    </thead>\n'
        html_template += '                    <tbody>\n'
        
        # 都道府県データ行 (Prefecture data rows)
        for pref_name, pref_data in tab_data['prefectures'].items():
            html_template += '                        <tr>\n'
            html_template += f'                            <td>{pref_name}</td>\n'
            
            for j, value in enumerate(pref_data):
                # 最後の列が計列かチェック
                if j == len(pref_data) - 1:
                    html_template += f'                            <td class="total-col">{value}</td>\n'
                else:
                    html_template += f'                            <td>{value}</td>\n'
            
            html_template += '                        </tr>\n'
        
        # 総計行 (Total row)
        if tab_data.get('total'):
            html_template += '                        <tr class="total-row">\n'
            html_template += '                            <td>総計</td>\n'
            
            for j, value in enumerate(tab_data['total']):
                if j == len(tab_data['total']) - 1:
                    html_template += f'                            <td class="total-col">{value}</td>\n'
                else:
                    html_template += f'                            <td>{value}</td>\n'
            
            html_template += '                        </tr>\n'
        
        html_template += '                    </tbody>\n'
        html_template += '                </table>\n'
        html_template += '            </div>\n'
        html_template += '        </div>\n\n'
    
    # JavaScriptを追加 (Add JavaScript)
    html_template += '''        <script>
            // タブ切り替え関数 (Tab switching function)
            function showTab(evt, tabName) {
                var i, tabcontent, tabbuttons;
                
                // すべてのタブコンテンツを非表示 (Hide all tab content)
                tabcontent = document.getElementsByClassName("tab-content");
                for (i = 0; i < tabcontent.length; i++) {
                    tabcontent[i].classList.remove("active");
                }
                
                // すべてのタブボタンを非アクティブ (Deactivate all tab buttons)
                tabbuttons = document.getElementsByClassName("tab");
                for (i = 0; i < tabbuttons.length; i++) {
                    tabbuttons[i].classList.remove("active");
                }
                
                // 選択されたタブを表示 (Show selected tab)
                document.getElementById(tabName).classList.add("active");
                evt.currentTarget.classList.add("active");
                
                // 合計列までスクロール (Scroll to total column)
                setTimeout(function() {
                    var tableWrapper = document.querySelector("#" + tabName + " .table-wrapper");
                    if (tableWrapper) {
                        tableWrapper.scrollLeft = tableWrapper.scrollWidth;
                    }
                }, 100);
            }
            
            // ページロード時に最初のタブの合計列までスクロール (Scroll to total column on page load)
            window.addEventListener('load', function() {
                var firstTabContent = document.querySelector('.tab-content.active .table-wrapper');
                if (firstTabContent) {
                    firstTabContent.scrollLeft = firstTabContent.scrollWidth;
                }
            });
            
            // クロスページナビゲーションを設定 (Set up cross-page navigation)
            var currentPage = window.location.pathname;
            var crossNav = document.getElementById('cross-nav');
            
            if (currentPage.includes('koufu.html') || currentPage.endsWith('/')) {
                crossNav.innerHTML = '<a href="koufu3.html">充電設備</a> | <a href="koufu2.html">外部給電器（V2L）･V2H充放電設備</a>';
            } else if (currentPage.includes('koufu2.html')) {
                crossNav.innerHTML = '<a href="koufu.html">EV・PHEV・FCV・原付EV</a> | <a href="koufu3.html">充電設備</a>';
            } else if (currentPage.includes('koufu3.html')) {
                crossNav.innerHTML = '<a href="koufu.html">EV・PHEV・FCV・原付EV</a> | <a href="koufu2.html">外部給電器（V2L）･V2H充放電設備</a>';
            }
        </script>
    </div>
</body>
</html>
'''
    
    # HTMLファイルを書き込み (Write HTML file)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_template)
    
    print(f"✅ HTMLファイルを生成しました: {output_file}")
    print(f"   Generated HTML file: {output_file}")


def main():
    if len(sys.argv) < 3:
        print("使い方 (Usage):")
        print("  python excel_to_html_converter.py <excel_file> <output_file>")
        print("\n例 (Examples):")
        print("  python excel_to_html_converter.py koufu_data.xlsx koufu.html")
        print("  python excel_to_html_converter.py koufu2_data.xlsx koufu2.html")
        print("  python excel_to_html_converter.py koufu3_data.xlsx koufu3.html")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # 出力ファイル名からページタイトルを決定 (Determine page title from output filename)
    if 'koufu2' in output_file:
        page_title = "都道府県別補助金交付状況"
        page_subtitle = "外部給電器（V2L）･V2H充放電設備"
    elif 'koufu3' in output_file:
        page_title = "都道府県別補助金交付状況"
        page_subtitle = "充電設備"
    else:
        page_title = "都道府県別補助金交付状況"
        page_subtitle = "EV・PHEV・FCV・原付EV"
    
    # ファイル存在チェック (Check if file exists)
    if not Path(excel_file).exists():
        print(f"❌ エラー: ファイルが見つかりません: {excel_file}")
        print(f"   Error: File not found: {excel_file}")
        sys.exit(1)
    
    print(f"📖 エクセルファイルを読み込んでいます: {excel_file}")
    print(f"   Reading Excel file: {excel_file}")
    
    # データを読み込み (Read data)
    data = read_excel_data(excel_file)
    
    print(f"📊 {len(data)}個のシートを読み込みました")
    print(f"   Loaded {len(data)} sheets")
    for sheet_name in data.keys():
        print(f"   - {sheet_name}")
    
    # HTMLを生成 (Generate HTML)
    print(f"\n🔨 HTMLを生成しています...")
    print(f"   Generating HTML...")
    generate_html(data, output_file, page_title, page_subtitle)
    
    print(f"\n✨ 完了しました!")
    print(f"   Done!")


if __name__ == '__main__':
    main()
